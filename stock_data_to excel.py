import os
from openpyxl import load_workbook
import pandas as pd

# 定义目标文件夹路径和 Excel 文件路径
target_folder = 'D:/documents/GitHub/stock_analysis/results'  # 替换为你的目标文件夹路径
excel_file = 'D:/bigdata/all_stocks_analysis_0323_demo.xlsx'  # 替换为你的 Excel 文件路径

# 检查文件是否存在
if os.path.exists(excel_file):
    try:
        # 加载 Excel 文件
        wb = load_workbook(excel_file)
        print("Excel 文件加载成功")

        # 定义 results_df（假设需要存储股票分析结果）
        results_df = pd.DataFrame(columns=["股票代码", "市场类型", "历史波动率", "模型选择"])

        # 保存结果到文件
        results_dir = 'results'  # 修改为当前目录下的 results 文件夹
        os.makedirs(results_dir, exist_ok=True)  # 创建目录（如果不存在）
        output_file_path = os.path.join(results_dir, 'all_stocks_analysis.csv')
        try:
            # 保存为 CSV 文件
            results_df.to_csv(output_file_path, index=False, encoding='utf-8-sig')
            print(f"分析结果已保存至: {output_file_path}")
        except PermissionError:
            print(f"错误：无法保存文件到 {output_file_path}，请检查目录权限")

        # 加载 Excel 文件
        wb = load_workbook(excel_file)
        ws = wb.active

        # 遍历 Excel 文件中的每一行，从第二行开始（假设第一行是标题）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            stock_code = row[0]  # 假设股票代码在第一列
            if not stock_code:
                print(f"警告：第 {row_idx} 行股票代码为空，跳过")
                continue

            # 构建对应的 .txt 文件路径
            txt_file = None
            for file in os.listdir(target_folder):
                if file.startswith(str(stock_code)) and file.endswith('.txt'):
                    txt_file = os.path.join(target_folder, file)
                    break

            if not txt_file:
                print(f"未找到股票代码 {stock_code} 对应的 .txt 文件")
                continue

            # 从 .txt 文件中读取所有信息
            try:
                with open(txt_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()

                # 解析 .txt 文件内容
                data = {}
                for line in lines:
                    if ':' in line:  # 假设每行数据是以冒号分隔的键值对
                        key, value = line.strip().split(':', 1)  # 只分割第一个冒号
                        data[key.strip()] = value.strip()

                # 打印解析后的数据（调试用）
                print(f"解析后的数据（股票代码 {stock_code}）：", data)

                # 定义 Excel 文件的列映射（根据你的需求调整）
                column_mapping = {
                    "股票代码": "A",  # 股票代码填充到 A 列
                    "股票名称": "B",  # 股票名称填充到 B 列
                    "均方根误差(RMSE)": "F",  # RMSE 填充到 F 列
                    "平均百分比误差(MAPE)": "G",  # MAPE 填充到 G 列
                    "2025-03-25": "H",  # 2025-03-25 预测价格填充到 H 列
                    "2025-03-26": "I",  # 2025-03-26 预测价格填充到 I 列
                    "2025-03-27": "J",  # 2025-03-27 预测价格填充到 J 列
                    "2025-03-28": "K",  # 2025-03-28 预测价格填充到 K 列
                    "2025-03-31": "L",  # 2025-03-31 预测价格填充到 L 列
                    "预测5日涨跌幅": "M",  # 预测5日涨跌幅填充到 M 列
                    "当前趋势": "N",  # 当前趋势填充到 N 列
                    "预测趋势": "O",  # 预测趋势填充到 O 列
                }

                # 将解析后的数据填充到 Excel 文件
                for key, value in data.items():
                    if key in column_mapping:
                        column = column_mapping[key]
                        # 处理数值和字符串
                        if key in ["均方根误差(RMSE)", "平均百分比误差(MAPE)", "预测5日涨跌幅"]:
                            # 去掉单位（如“元”或“%”），只保留数值
                            value = value.replace("元", "").replace("%", "").strip()
                        ws[f"{column}{row_idx}"] = value

            except (UnicodeDecodeError, ValueError, IndexError) as e:
                print(f"错误：文件 {txt_file} 内容格式不正确或编码错误: {str(e)}")
                continue

        # 保存修改后的 Excel 文件
        wb.save(excel_file)
        print("Excel 文件已更新并保存")

    except Exception as e:
        print(f"加载 Excel 文件时出错: {str(e)}")
else:
    print(f"错误：文件 {excel_file} 不存在，请检查路径和文件名")